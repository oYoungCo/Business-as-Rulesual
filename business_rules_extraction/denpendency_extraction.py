import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openai import OpenAI
from tqdm import tqdm
from datetime import datetime
from dotenv import load_dotenv

# ================= Configuration =================
# Load environment variables (API keys) from .env file
load_dotenv()

CONFIG = {
    # Input Excel file containing the dependency analysis prompts
    "INPUT_FILE": "",
    
    # Column name in the input Excel that contains the actual prompt text
    "PROMPT_COL": "",
    
    # Output file path for results
    "OUTPUT_FILE": "",
    
    # Model identifier
    "MODEL_NAME": "",
    
    # API Key from environment variables
    "API_KEY": os.getenv("DASHSCOPE_API_KEY"),
    
    # Base URL for the API service
    "BASE_URL": "https://dashscope.aliyuncs.com/compatible-mode/v1",
    
    # Sampling temperature (lower = more deterministic)
    "TEMPERATURE": 0.01,
    
    # Max retries for network errors
    "MAX_RETRIES": 3,
    
    # Delay between retries (seconds)
    "RETRY_DELAY": 2
}

# ================= Helper Functions =================

def get_completion(client, prompt):
    """
    Send request to LLM with retry mechanism.
    """
    for attempt in range(CONFIG["MAX_RETRIES"]):
        try:
            completion = client.chat.completions.create(
                model=CONFIG["MODEL_NAME"],
                messages=[
                    {'role': 'system', 'content': 'You are a helpful assistant.'},
                    {'role': 'user', 'content': prompt}
                ],
                temperature=CONFIG["TEMPERATURE"],
            )
            return completion.choices[0].message.content
        except Exception as e:
            if attempt < CONFIG["MAX_RETRIES"] - 1:
                print(f"\n[Warning] API request failed: {e}. Retrying in {CONFIG['RETRY_DELAY']}s...")
                time.sleep(CONFIG["RETRY_DELAY"])
            else:
                return f"ERROR: {str(e)}"

def initialize_output_file(filepath):
    """
    Initialize output Excel file. 
    Returns the number of already processed rows for resume capability.
    """
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dependency_Results"
        ws.append(["Index", "Original Prompt", CONFIG["MODEL_NAME"], "Timestamp"])
        wb.save(filepath)
        return 0
    else:
        # If file exists, count rows to determine where to resume
        try:
            df_exist = pd.read_excel(filepath)
            return len(df_exist)
        except Exception:
            return 0

def save_result(filepath, index, prompt, response, timestamp):
    """
    Append a single result row to the Excel file.
    """
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append([index, prompt, response, timestamp])
        wb.save(filepath)
    except Exception as e:
        print(f"\n[Error] Failed to write to Excel: {e}")

# ================= Main Execution =================

def main():
    # 1. Validate API Key
    if not CONFIG["API_KEY"]:
        raise ValueError("API Key not found! Please set 'DASHSCOPE_API_KEY' in your .env file.")

    client = OpenAI(
        api_key=CONFIG["API_KEY"],
        base_url=CONFIG["BASE_URL"],
    )

    # 2. Load Data
    print(f"Loading input file: {CONFIG['INPUT_FILE']}...")
    try:
        df = pd.read_excel(CONFIG['INPUT_FILE'])
        df.dropna(how='all', inplace=True)
        df.columns = df.columns.str.strip()
        
        if CONFIG["PROMPT_COL"] not in df.columns:
            raise ValueError(f"Column '{CONFIG['PROMPT_COL']}' not found. Available columns: {list(df.columns)}")
            
        prompt_list = df[CONFIG["PROMPT_COL"]].tolist()
    except Exception as e:
        print(f"Error reading input file: {e}")
        return

    # 3. Initialize & Resume Logic
    processed_count = initialize_output_file(CONFIG["OUTPUT_FILE"])
    start_index = processed_count
    
    print(f"Total prompts: {len(prompt_list)}")
    print(f"Already processed: {processed_count}")
    print(f"Starting from index: {start_index + 1}")

    if start_index >= len(prompt_list):
        print("All prompts have been processed!")
        return

    # 4. Processing Loop
    for i, prompt in enumerate(tqdm(prompt_list[start_index:], desc="Extracting Dependencies", initial=start_index, total=len(prompt_list))):
        
        real_index = start_index + i + 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # API Call
        response_text = get_completion(client, prompt)
        
        # Save Result
        save_result(CONFIG["OUTPUT_FILE"], real_index, prompt, response_text, timestamp)

    print("\nProcessing Finished Successfully!")

if __name__ == "__main__":
    main()
