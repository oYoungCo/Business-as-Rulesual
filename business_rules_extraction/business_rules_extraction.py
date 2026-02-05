#!/usr/bin/env bash 
import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openai import OpenAI
from tqdm import tqdm
from datetime import datetime
from dotenv import load_dotenv  # 需要 pip install python-dotenv

# ================= Configuration =================
# Load environment variables from the .env file
load_dotenv()

CONFIG = {
    # Path to the input Excel file containing prompts
    "INPUT_FILE": "",
    
    # Path where the extraction results will be saved
    "OUTPUT_FILE": "",
    
    # The specific column name in the Excel file that contains the prompts
    "PROMPT_COL": "",
    
    # The model identifier to be used for API calls
    "MODEL_NAME": "",
    
    # API Key retrieved securely from environment variables
    "API_KEY": os.getenv("DASHSCOPE_API_KEY"),
    
    # The base URL for the API endpoint (e.g., for Alibaba DashScope/DeepSeek)
    "BASE_URL": "https://dashscope.aliyuncs.com/compatible-mode/v1",
    
    # Sampling temperature: lower values (e.g., 0.01) make output more deterministic
    "TEMPERATURE": 0.01,
    
    # Maximum number of retry attempts for failed API requests
    "MAX_RETRIES": 3,
    
    # Delay in seconds to wait before retrying a failed request
    "RETRY_DELAY": 2
}



def get_completion(client, prompt):
   
    for attempt in range(CONFIG["MAX_RETRIES"]):
        try:
            completion = client.chat.completions.create(
                model=CONFIG["MODEL_NAME"],
                messages=[
                    {'role': 'system', 'content': 'You are a helpful assistant.'},
                    {'role': 'user', 'content': prompt}
                ],
                temperature=CONFIG["TEMPERATURE"],
            )
            return completion.choices[0].message.content
        except Exception as e:
            if attempt < CONFIG["MAX_RETRIES"] - 1:
                print(f"\n[Warning] API request failed: {e}. Retrying in {CONFIG['RETRY_DELAY']}s...")
                time.sleep(CONFIG["RETRY_DELAY"])
            else:
                return f"ERROR: {str(e)}"

def initialize_output_file(filepath):
    
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.append(["Index", "Original Prompt", CONFIG["MODEL_NAME"], "Timestamp"])
        wb.save(filepath)
        return 0
    else:
        try:
            df_exist = pd.read_excel(filepath)
            return len(df_exist)
        except Exception:
            return 0

def save_result(filepath, index, prompt, response, timestamp):
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append([index, prompt, response, timestamp])
        wb.save(filepath)
    except Exception as e:
        print(f"\n[Error] Failed to write to Excel: {e}")



def main():

    if not CONFIG["API_KEY"]:
        raise ValueError("API Key not found! Please set 'DASHSCOPE_API_KEY' in .env file or environment variables.")

    client = OpenAI(
        api_key=CONFIG["API_KEY"],
        base_url=CONFIG["BASE_URL"],
    )

    print(f"Loading input file: {CONFIG['INPUT_FILE']}...")
    try:
        df = pd.read_excel(CONFIG['INPUT_FILE'])
        df.dropna(how='all', inplace=True) 
        df.columns = df.columns.str.strip()
        
        if CONFIG["PROMPT_COL"] not in df.columns:
            raise ValueError(f"Column '{CONFIG['PROMPT_COL']}' not found in Excel.")
            
        prompt_list = df[CONFIG["PROMPT_COL"]].tolist()
    except Exception as e:
        print(f"Error reading input file: {e}")
        return

    processed_count = initialize_output_file(CONFIG["OUTPUT_FILE"])
    start_index = processed_count
    
    print(f"Total prompts: {len(prompt_list)}")
    print(f"Already processed: {processed_count}")
    print(f"Starting from index: {start_index + 1}")

    if start_index >= len(prompt_list):
        print("All prompts have been processed!")
        return


    for i, prompt in enumerate(tqdm(prompt_list[start_index:], desc="Processing", initial=start_index, total=len(prompt_list))):
        
        real_index = start_index + i + 1  
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        response_text = get_completion(client, prompt)

        save_result(CONFIG["OUTPUT_FILE"], real_index, prompt, response_text, timestamp)

    print("\nProcessing Finished Successfully!")

if __name__ == "__main__":
    main()
